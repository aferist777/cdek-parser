"""CSV and XLSX exports of whatever the list is currently showing."""
from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import query

HEADERS = [
    ("discount", "Скидка, %"),
    ("price", "Цена"),
    ("old_price", "Старая цена"),
    ("title", "Товар"),
    ("brand", "Бренд"),
    ("root", "Раздел"),
    ("badge", "Метка"),
    ("url", "Ссылка"),
    ("product_id", "ID"),
    ("last_seen", "Обновлено"),
]


def _row_values(row: dict, root_titles: dict) -> list:
    """One export row: readable section name, local timestamp."""
    values = []
    for key, _ in HEADERS:
        value = row.get(key)
        if key == "root":
            value = root_titles.get(value, value)
        elif key == "last_seen" and value:
            try:
                value = datetime.fromisoformat(value).astimezone().strftime("%Y-%m-%d %H:%M")
            except ValueError:
                pass
        values.append(value)
    return values


def _stamp(prefix: str, suffix: str, out_dir: str | Path) -> Path:
    directory = Path(out_dir)
    directory.mkdir(parents=True, exist_ok=True)
    return directory / f"{prefix}-{datetime.now().strftime('%Y%m%d-%H%M')}.{suffix}"


def to_csv(store, filters: dict, sort: str, out_dir: str | Path, root_titles: dict | None = None) -> Path:
    root_titles = root_titles or {}
    path = _stamp("cdek-sale", "csv", out_dir)
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow([title for _, title in HEADERS])
        for row in query.iterate(store, filters, sort):
            writer.writerow(_row_values(row, root_titles))
    return path


def to_xlsx(store, filters: dict, sort: str, out_dir: str | Path, root_titles: dict | None = None) -> Path:
    root_titles = root_titles or {}
    path = _stamp("cdek-sale", "xlsx", out_dir)
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Скидки"

    ws.append([title for _, title in HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in query.iterate(store, filters, sort):
        ws.append(_row_values(row, root_titles))

    widths = {"discount": 10, "price": 12, "old_price": 13, "title": 60, "brand": 20,
              "root": 12, "badge": 10, "url": 46, "product_id": 11, "last_seen": 21}
    for index, (key, _) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(key, 14)

    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path
